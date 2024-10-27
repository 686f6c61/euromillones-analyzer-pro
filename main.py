import sys
import os
from datetime import datetime
from modules.data_loader import DataLoader
from modules.analyzer import DataAnalyzer
from modules.predictor import Predictor
from modules.generator import CombinationGenerator
from modules.statistics import Statistics
from modules.visualizer import Visualizer
from modules.helpers import Helpers
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class EuromillonesApp:
    def __init__(self):
        self.helper = Helpers()
        self.loader = DataLoader()
        self.df = self.loader.load_data()
        self.analyzer = DataAnalyzer(self.df)
        self.predictor = Predictor(self.df)
        self.generator = CombinationGenerator(self.df)
        self.statistics = Statistics(self.df)
        self.visualizer = Visualizer()

    def run(self):
        """Ejecuta la aplicación principal."""
        while True:
            self.show_main_menu()
            option = input("\nSeleccione una opción: ")

            if option == "1":
                self.quick_analysis()
            elif option == "2":
                self.detailed_analysis()
            elif option == "3":
                self.predictions_menu()
            elif option == "4":
                self.combination_generator()
            elif option == "5":
                self.statistics_menu()
            elif option == "6":
                self.export_data()
            elif option == "7":
                print("\n👋 ¡Hasta luego!")
                break
            else:
                print("\n❌ Opción no válida. Por favor, intente de nuevo.")

            input("\nPresione Enter para continuar...")

    def show_main_menu(self):
        """Muestra el menú principal."""
        print("\n=== 🎲 EUROMILLONES ANALYZER PRO 🎲 ===")
        print("\n1. Análisis Rápido")
        print("2. Análisis Detallado")
        print("3. Predicciones")
        print("4. Generador de Combinaciones")
        print("5. Estadísticas")
        print("6. Exportar Datos")
        print("7. Salir")

    def quick_analysis(self):
        """Menú de análisis rápido."""
        while True:
            self.helper.clear_screen()
            print("\n=== ⚡ ANLISIS RÁPIDO ===")
            print("\n1. Últimos 5 sorteos")
            print("2. Números más frecuentes")
            print("3. Analizar mi combinación")
            print("4. Estadísticas rápidas")
            print("5. Volver")

            option = input("\nSeleccione una opción: ")

            if option == "1":
                self.statistics.show_last_draws(5)
            elif option == "2":
                self.statistics.show_frequent_numbers()
            elif option == "3":
                self.analyzer.analyze_user_combination()
            elif option == "4":
                self.statistics.quick_stats()
            elif option == "5":
                break
            else:
                print("Opción no válida. Intente de nuevo.")

            input("\nPresione Enter para continuar...")

    def detailed_analysis(self):
        """Menú de análisis detallado."""
        while True:
            self.helper.clear_screen()
            print("\n=== 🔍 ANÁLISIS DETALLADO ===")
            print("\n1. Análisis Temporal")
            print("2. Análisis de Patrones")
            print("3. Análisis de Correlaciones")
            print("4. Análisis de Estrellas")
            print("5. Análisis Personalizado")
            print("6. Volver")

            option = input("\nSeleccione una opción: ")

            if option == "1":
                self.statistics.temporal_analysis()
            elif option == "2":
                self.statistics.pattern_analysis()
            elif option == "3":
                self.statistics.analyze_correlations()
            elif option == "4":
                self.analyzer.analyze_stars()
            elif option == "5":
                self.custom_analysis_menu()
            elif option == "6":
                break
            else:
                print("Opción no válida. Intente de nuevo.")

            input("\nPresione Enter para continuar...")

    def temporal_analysis_menu(self):
        """Menú de análisis temporal."""
        while True:
            self.helper.clear_screen()
            print("\n=== 📅 ANÁLISIS TEMPORAL ===")
            print("\n1. Tendencias mensuales")
            print("2. Análisis por estaciones")
            print("3. Comparación año actual vs anterior")
            print("4. Análisis de días de la semana")
            print("5. Volver")

            option = input("\nSeleccione una opción: ")

            if option == "1":
                self.analyzer.analyze_monthly_trends()
            elif option == "2":
                self.analyzer.analyze_seasonal_trends()
            elif option == "3":
                self.analyzer.compare_current_vs_previous_year()
            elif option == "4":
                self.analyzer.analyze_weekday_trends()
            elif option == "5":
                break
            else:
                input("\n❌ Opción no válida. Presione Enter para continuar...")

            input("\nPresione Enter para continuar...")

    def predictions_menu(self):
        """Menú de predicciones."""
        while True:
            self.helper.clear_screen()
            print("\n=== 🔮 PREDICCIONES ===")
            print("\n1. Números calientes")
            print("2. Números fríos")
            print("3. Predicción próximo sorteo")
            print("4. Análisis probabilístico")
            print("5. Volver")

            option = input("\nSeleccione una opción: ")

            if option == "1":
                self.predictor.get_hot_numbers()
            elif option == "2":
                self.predictor.get_cold_numbers()
            elif option == "3":
                self.predictor.predict_next_draw()
            elif option == "4":
                self.statistics.get_probability_analysis()
            elif option == "5":
                break

            input("\nPresione Enter para continuar...")

    def combination_generator(self):
        """Menú del generador de combinaciones."""
        while True:
            self.helper.clear_screen()
            print("\n=== 🎲 GENERADOR DE COMBINACIONES ===")
            print("\n1. Combinación estadística")
            print("2. Combinación equilibrada")
            print("3. Combinación personalizada")
            print("4. Múltiples combinaciones")
            print("5. Volver")

            option = input("\nSeleccione una opción: ")

            if option == "1":
                self.generate_statistical_combination()
            elif option == "2":
                self.generate_balanced_combination()
            elif option == "3":
                self.generate_custom_combination()
            elif option == "4":
                self.generate_multiple_combinations()
            elif option == "5":
                break

            input("\nPresione Enter para continuar...")

    def statistics_menu(self):
        """Menú de estadísticas."""
        while True:
            self.helper.clear_screen()
            print("\n=== 📊 ESTADÍSTICAS ===")
            print("1. Estadísticas básicas")
            print("2. Estadísticas avanzadas")
            print("3. Análisis estacional")
            print("4. Análisis de rachas")
            print("5. Informe completo")
            print("6. Números más frecuentes y combinaciones")
            print("7. Volver")

            option = input("\nSeleccione una opción: ")

            if option == "1":
                self.statistics.show_basic_stats()
            elif option == "2":
                self.statistics.get_advanced_stats()
            elif option == "3":
                self.statistics.get_seasonal_analysis()
            elif option == "4":
                self.statistics.streak_analysis()
            elif option == "5":
                self.statistics.get_complete_report()
            elif option == "6":
                self.statistics.show_frequent_numbers()
            elif option == "7":
                break
            else:
                print("Opción no válida. Intente de nuevo.")

            input("\nPresione Enter para continuar...")

    def export_data(self):
        """Menú de exportación de datos."""
        while True:
            self.helper.clear_screen()
            print("\n=== 💾 EXPORTAR DATOS ===")
            print("\n1. Exportar a CSV")
            print("2. Exportar a Excel")
            print("3. Exportar a JSON")
            print("4. Volver")

            option = input("\nSeleccione una opción: ")

            if option == "1":
                filename = input("Nombre del archivo (sin extensión): ") + ".csv"
                success, message = self.loader.export_data(filename, 'csv')
            elif option == "2":
                filename = input("Nombre del archivo (sin extensión): ") + ".xlsx"
                success, message = self.loader.export_data(filename, 'excel')
            elif option == "3":
                filename = input("Nombre del archivo (sin extensión): ") + ".json"
                success, message = self.loader.export_data(filename, 'json')
            elif option == "4":
                break

            if 'success' in locals():
                print(f"\n{'✅' if success else '❌'} {message}")
            input("\nPresione Enter para continuar...")

    def analyze_my_combination(self):
        """Analiza una combinación ingresada por el usuario."""
        print("\n🔍 ANÁLISIS DE COMBINACIÓN")
        try:
            numbers = []
            print("\nIngrese los números (1-50):")
            for i in range(5):
                while True:
                    num = int(input(f"Número {i+1}: "))
                    valid, message = self.helper.validate_number_range(num, 1, 50)
                    if valid and num not in numbers:
                        numbers.append(num)
                        break
                    print(f"❌ {message}")

            stars = []
            print("\nIngrese las estrellas (1-12):")
            for i in range(2):
                while True:
                    star = int(input(f"Estrella {i+1}: "))
                    valid, message = self.helper.validate_number_range(star, 1, 12)
                    if valid and star not in stars:
                        stars.append(star)
                        break
                    print(f"❌ {message}")

            self.analyzer.analyze_combination(numbers, stars)
            
        except ValueError:
            print("\n❌ Error: Debe introducir números válidos")

    def custom_analysis_menu(self):
        """Menú de análisis personalizado."""
        while True:
            print("\n=== 🎯 ANÁLISIS PERSONALIZADO ===")
            print("1. Análisis por rango de fechas")
            print("2. Análisis por número específico")
            print("3. Búsqueda por patrón")
            print("4. Volver")

            option = input("\nSeleccione una opción: ")

            if option == "1":
                self.date_range_analysis()
            elif option == "2":
                self.single_number_analysis()
            elif option == "3":
                self.pattern_search()
            elif option == "4":
                break
            else:
                print("Opción no válida. Intente de nuevo.")

            input("\nPresione Enter para continuar...")

    def date_range_analysis(self):
        """Análisis por rango de fechas."""
        print("\n📅 ANÁLISIS POR RANGO DE FECHAS")
        
        # Obtener fechas de inicio y fin
        start_date = input("Ingrese la fecha de inicio (DD-MM-YYYY): ")
        end_date = input("Ingrese la fecha de fin (DD-MM-YYYY): ")
        
        try:
            start_date = datetime.strptime(start_date, "%d-%m-%Y")
            end_date = datetime.strptime(end_date, "%d-%m-%Y")
        except ValueError:
            print("❌ Formato de fecha inválido. Use DD-MM-YYYY.")
            return
        
        # Filtrar el DataFrame por el rango de fechas
        mask = (self.loader.df['fecha'] >= start_date) & (self.loader.df['fecha'] <= end_date)
        df_filtered = self.loader.df.loc[mask]
        
        if df_filtered.empty:
            print("❌ No hay datos para el rango de fechas especificado.")
            return
        
        # Realizar análisis en el rango de fechas
        print(f"\nAnálisis para el período {start_date.date()} a {end_date.date()}:")
        print(f"Número de sorteos: {len(df_filtered)}")
        
        # Análisis de números más frecuentes
        numeros = []
        for col in ['n1', 'n2', 'n3', 'n4', 'n5']:
            numeros.extend(df_filtered[col].tolist())
        
        freq = pd.Series(numeros).value_counts().head(10)
        print("\nNúmeros más frecuentes:")
        for num, count in freq.items():
            print(f"Número {num}: {count} apariciones")
        
        # Análisis de estrellas más frecuentes
        estrellas = []
        for col in ['e1', 'e2']:
            estrellas.extend(df_filtered[col].tolist())
        
        freq_estrellas = pd.Series(estrellas).value_counts().head(5)
        print("\nEstrellas más frecuentes:")
        for star, count in freq_estrellas.items():
            print(f"Estrella {star}: {count} apariciones")

    def single_number_analysis(self):
        """Análisis por número específico."""
        print("\n ANÁLISIS POR NÚMERO ESPECÍFICO")
        
        number = Helpers.get_valid_input(
            "Ingrese un número (1-50): ", 
            int, 
            lambda x: 1 <= x <= 50
        )
        
        # Análisis del número
        apariciones = sum((self.loader.df[col] == number).sum() for col in ['n1', 'n2', 'n3', 'n4', 'n5'])
        porcentaje = (apariciones / len(self.loader.df)) * 100
        
        print(f"\nAnálisis para el número {number}:")
        print(f"Apariciones totales: {apariciones}")
        print(f"Porcentaje de sorteos: {porcentaje:.2f}%")
        
        # Última aparición
        ultima_aparicion = None
        for _, row in self.loader.df.iterrows():
            if number in [row['n1'], row['n2'], row['n3'], row['n4'], row['n5']]:
                ultima_aparicion = row['fecha']
                break
        
        if ultima_aparicion:
            dias_desde_ultima = (datetime.now().date() - ultima_aparicion.date()).days
            print(f"Última aparición: {ultima_aparicion.strftime('%d-%m-%Y')} ({dias_desde_ultima} días atrás)")
        else:
            print("Este número nunca ha salido.")

    def pattern_search(self):
        """Búsqueda por patrón."""
        print("\n🔍 BÚSQUEDA POR PATRÓN")
        
        print("Ingrese el patrón de búsqueda (use 'X' para cualquier número):")
        pattern = []
        for i in range(5):
            while True:
                value = input(f"Posición {i+1}: ").upper()
                if value == 'X' or (value.isdigit() and 1 <= int(value) <= 50):
                    pattern.append(value)
                    break
                print("Entrada inválida. Use 'X' o un nmero entre 1 y 50.")
        
        # Buscar coincidencias
        coincidencias = []
        for _, row in self.loader.df.iterrows():
            if all((p == 'X' or int(p) == row[f'n{i+1}']) for i, p in enumerate(pattern)):
                coincidencias.append(row)
        
        if coincidencias:
            print(f"\nSe encontraron {len(coincidencias)} coincidencias:")
            for row in coincidencias:
                numeros = ' - '.join(map(str, [row['n1'], row['n2'], row['n3'], row['n4'], row['n5']]))
                estrellas = f"{row['e1']} - {row['e2']}"
                print(f"{row['fecha'].strftime('%d-%m-%Y')}: {numeros} | Estrellas: {estrellas}")
        else:
            print("No se encontraron coincidencias para el patrón especificado.")

    def generate_statistical_combination(self):
        combinations = self.generator.generate_statistical(num_combinations=1)
        self.generator.evaluate_combinations(combinations)

    def generate_balanced_combination(self):
        combinations = self.generator.generate_balanced(num_combinations=1)
        self.generator.evaluate_combinations(combinations)

    def generate_custom_combination(self):
        # Implementar lógica para obtener números forzados y prohibidos del usuario
        forced_numbers = []
        forbidden_numbers = []
        combinations = self.generator.generate_custom(forced_numbers, forbidden_numbers, num_combinations=1)
        self.generator.evaluate_combinations(combinations)

    def generate_multiple_combinations(self):
        num_combinations = int(input("¿Cuántas combinaciones desea generar? "))
        combinations = self.generator.generate_smart(num_combinations=num_combinations)
        self.generator.evaluate_combinations(combinations)

    def generate_full_report(self):
        print("\n📊 GENERANDO INFORME COMPLETO")
        
        wb = Workbook()
        
        try:
            self._generate_basic_stats(wb)
            print("Estadísticas básicas generadas")
        except Exception as e:
            print(f"Error al generar estadísticas básicas: {str(e)}")
        
        try:
            self._generate_number_frequency(wb)
            print("Análisis de frecuencia de números generado")
        except Exception as e:
            print(f"Error al generar frecuencia de números: {str(e)}")
        
        try:
            self._generate_star_frequency(wb)
            print("Análisis de frecuencia de estrellas generado")
        except Exception as e:
            print(f"Error al generar frecuencia de estrellas: {str(e)}")
        
        try:
            self._generate_hot_cold_numbers(wb)
            print("Análisis de números calientes y fríos generado")
        except Exception as e:
            print(f"Error al generar números calientes y fríos: {str(e)}")
        
        try:
            self._generate_predictions(wb)
            print("Predicciones generadas")
        except Exception as e:
            print(f"Error al generar predicciones: {str(e)}")
        
        try:
            self._generate_patterns(wb)
            print("Análisis de patrones generado")
        except Exception as e:
            print(f"Error al generar análisis de patrones: {str(e)}")
        
        try:
            self._generate_combination_analysis(wb)
            print("Análisis de combinaciones generado")
        except Exception as e:
            print(f"Error al generar análisis de combinaciones: {str(e)}")
        
        filename = "Informe_Euromillones.xlsx"
        wb.save(filename)
        print(f"\nInforme completo generado y guardado como {filename}")

    def _generate_basic_stats(self, wb):
        ws = wb.active
        ws.title = "Estadísticas Básicas"
        ws.append(["Estadísticas Básicas del Euromillones"])
        ws.append([])
        
        stats = self.statistics.show_basic_stats()
        for stat in stats:
            ws.append(stat)
        
        self._format_worksheet(ws)

    def _generate_number_frequency(self, wb):
        ws = wb.create_sheet("Frecuencia de Números")
        ws.append(["Frecuencia de Números"])
        ws.append([])
        
        freq_data = self.statistics.number_frequency_analysis()
        ws.append(["Número", "Frecuencia", "Porcentaje", "Desviación"])
        for row in freq_data:
            ws.append(row)
        
        self._format_worksheet(ws)

    def _generate_star_frequency(self, wb):
        ws = wb.create_sheet("Frecuencia de Estrellas")
        ws.append(["Frecuencia de Estrellas"])
        ws.append([])
        
        star_data = self.statistics.star_frequency_analysis()
        ws.append(["Estrella", "Frecuencia", "Porcentaje", "Desviación"])
        for row in star_data:
            ws.append(row)
        
        self._format_worksheet(ws)

    def _generate_hot_cold_numbers(self, wb):
        ws = wb.create_sheet("Números Calientes y Fríos")
        ws.append(["Números Calientes y Fríos"])
        ws.append([])
        
        hot_numbers = self.predictor.get_hot_numbers()
        cold_numbers = self.predictor.get_cold_numbers()
        
        ws.append(["Números Calientes"])
        ws.append(["Número", "Frecuencia", "Porcentaje", "Tendencia"])
        for row in hot_numbers:
            ws.append(row)
        
        ws.append([])
        ws.append(["Números Fríos"])
        ws.append(["Número", "Frecuencia", "Porcentaje", "Tendencia"])
        for row in cold_numbers:
            ws.append(row)
        
        self._format_worksheet(ws)

    def _generate_predictions(self, wb):
        ws = wb.create_sheet("Predicciones")
        ws.append(["Predicciones para el Próximo Sorteo"])
        ws.append([])
        
        prediction = self.predictor.predict_next_draw()
        ws.append(["Números Predichos", "Estrellas Predichas"])
        ws.append([" - ".join(map(str, prediction['numbers'])), " - ".join(map(str, prediction['stars']))])
        
        ws.append([])
        ws.append(["Análisis de la Predicción"])
        for key, value in prediction['analysis'].items():
            ws.append([key, value])
        
        self._format_worksheet(ws)

    def _generate_patterns(self, wb):
        ws = wb.create_sheet("Patrones")
        ws.append(["Patrones Identificados"])
        ws.append([])
        
        patterns = self.analyzer.identify_patterns()
        for pattern_name, pattern_data in patterns.items():
            ws.append([pattern_name])
            for key, value in pattern_data.items():
                ws.append([key, value])
            ws.append([])
        
        self._format_worksheet(ws)

    def _generate_combination_analysis(self, wb):
        ws = wb.create_sheet("Análisis de Combinaciones")
        ws.append(["Análisis de Combinaciones"])
        ws.append([])
        
        combination_data = self.analyzer.analyze_combinations()
        for section, data in combination_data.items():
            ws.append([section])
            for key, value in data.items():
                ws.append([key, value])
            ws.append([])
        
        self._format_worksheet(ws)

    def _format_worksheet(self, ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

if __name__ == "__main__":
    try:
        app = EuromillonesApp()
        app.run()
    except KeyboardInterrupt:
        print("\n\n👋 ¡Hasta luego!")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error inesperado: {str(e)}")
        Helpers.log_error(e)
        sys.exit(1)
